import time
start_time = time.time()
def read_rows(file_name="practice", rows=2, start_column=1, end_column=2):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=file_name)
    workbook.sheetnames
    sheet = workbook.active

    lists = []
    n = start_column - 1
    while (n <= (end_column)):
        n += 1
        cell = (sheet.cell(row=rows, column=n))
        lists.append(cell.value)
    return lists


def read_table(file_name="practice", start_row=1, end_row=2, start_column=1, end_column=2):
    lists = []
    n = start_row - 1
    end_row -= 1
    while (n <= (end_row)):
        n += 1
        lists.append(read_rows(file_name, n, start_column, end_column))
    return lists


def protein_values(file_protei="file"):
    file_protein = file_protei + ".xlsx"
    return (read_table(file_protein, 24, 31, 3, 13))


def ABS_values(file_ab="file"):
    file_abs = file_ab + ".xlsx"
    return (read_table(file_abs, 24, 31, 3, 13))


def Atpase(abs=0, protein=1):
    x = (abs / 0.2431) * 1000
    w = x / protein
    v = w / 0.2
    o = v / 60
    Atpase = o * 4.5
    return Atpase


def carbonyl(abs=0):
    car = (abs * 0.18) / 132000  # car = carbonyl
    return car


def MDA(abs=0, protein=1):
    a = abs*3
    b = 1.56 * 100000 * 0.4 * protein
    MDA = a / b
    return MDA


def H2O2(abs=0, protein=1):
    a = 2.499 - abs
    b = 0.3175 * protein
    H2O2 = a / b
    return H2O2


def sulphy(abs=0):
    a = 1 - (abs / 14.150)
    b = (1.5 * a) / 1000
    c = (b * 1000) / 0.2
    return c


def column_result_Atpase(ABS_file="file", protein_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(Atpase((ABS_values()[r][n]), (protein_values()[r][n])))
        if (n == 11):
            break

    return lists


def table_result_Atpase(ABS_file="file", protein_file="file"):
    lists = []
    lists.append(column_result_Atpase(0))
    lists.append(column_result_Atpase(1))
    lists.append(column_result_Atpase(2))
    lists.append(column_result_Atpase(3))
    lists.append(column_result_Atpase(4))
    lists.append(column_result_Atpase(5))
    lists.append(column_result_Atpase(6))
    lists.append(column_result_Atpase(7))

    return lists


def column_result_sulphy(ABS_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(sulphy(ABS_values(ABS_file)[r][n]))
        if (n == 11):
            break

    return lists


def table_result_sulphy(ABS_file="file"):
    lists = []
    lists.append(column_result_sulphy(ABS_file, 0))
    lists.append(column_result_sulphy(ABS_file, 1))
    lists.append(column_result_sulphy(ABS_file, 2))
    lists.append(column_result_sulphy(ABS_file, 3))
    lists.append(column_result_sulphy(ABS_file, 4))
    lists.append(column_result_sulphy(ABS_file, 5))
    lists.append(column_result_sulphy(ABS_file, 6))
    lists.append(column_result_sulphy(ABS_file, 7))


def column_result_carbonyl(ABS_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(carbonyl(ABS_values(ABS_file)[r][n]))
        if (n == 11):
            break

    return lists


def table_result_carbonyl(ABS_file="file"):
    lists = []
    lists.append(column_result_carbonyl(ABS_file, 0))
    lists.append(column_result_carbonyl(ABS_file, 1))
    lists.append(column_result_carbonyl(ABS_file, 2))
    lists.append(column_result_carbonyl(ABS_file, 3))
    lists.append(column_result_carbonyl(ABS_file, 4))
    lists.append(column_result_carbonyl(ABS_file, 5))
    lists.append(column_result_carbonyl(ABS_file, 6))
    lists.append(column_result_carbonyl(ABS_file, 7))

    return lists


def column_result_H2O2(ABS_file="file", protein_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(H2O2((ABS_values(ABS_file)[r][n]), (protein_values(protein_file)[r][n])))
        if (n == 11):
            break

    return lists


def table_result_H2O2(ABS_file="file", protein_file="file"):
    lists = []
    lists.append(column_result_H2O2(ABS_file, protein_file, 0))
    lists.append(column_result_H2O2(ABS_file, protein_file, 1))
    lists.append(column_result_H2O2(ABS_file, protein_file, 2))
    lists.append(column_result_H2O2(ABS_file, protein_file, 3))
    lists.append(column_result_H2O2(ABS_file, protein_file, 4))
    lists.append(column_result_H2O2(ABS_file, protein_file, 5))
    lists.append(column_result_H2O2(ABS_file, protein_file, 6))
    lists.append(column_result_H2O2(ABS_file, protein_file, 7))

    return lists


def column_result_MDA(ABS_file="file", protein_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(MDA((ABS_values(ABS_file)[r][n]), (protein_values(protein_file)[r][n])))
        if (n == 11):
            break

    return lists


def table_result_MDA(ABS_file="file", protein_file="file"):
    lists = []
    lists.append(column_result_MDA(ABS_file, protein_file, 0))
    lists.append(column_result_MDA(ABS_file, protein_file, 1))
    lists.append(column_result_MDA(ABS_file, protein_file, 2))
    lists.append(column_result_MDA(ABS_file, protein_file, 3))
    lists.append(column_result_MDA(ABS_file, protein_file, 4))
    lists.append(column_result_MDA(ABS_file, protein_file, 5))
    lists.append(column_result_MDA(ABS_file, protein_file, 6))
    lists.append(column_result_MDA(ABS_file, protein_file, 7))

    return lists


def selection_4():
    ABS_file = input("Enter file containing ABS values: ")
    protein_file = input("Enter file containing Protein values: ")
    result = input("Save as: ")

    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    from openpyxl.chart import BarChart, Reference


    sheet["B24"] = "A"
    sheet["B25"] = "B"
    sheet["B26"] = "C"
    sheet["B27"] = "D"
    sheet["B28"] = "E"
    sheet["B29"] = "F"
    sheet["B30"] = "G"
    sheet["B31"] = "H"
    sheet["C23"] = "1"
    sheet["D23"] = "2"
    sheet["E23"] = "3"
    sheet["F23"] = "4"
    sheet["G23"] = "5"
    sheet["H23"] = "6"
    sheet["I23"] = "7"
    sheet["J23"] = "8"
    sheet["K23"] = "9"
    sheet["L23"] = "10"
    sheet["M23"] = "11"
    sheet["N23"] = "12"

    sheet["C24"] = table_result_MDA(ABS_file, protein_file)[0][0]
    sheet["D24"] = table_result_MDA(ABS_file, protein_file)[0][1]
    sheet["E24"] = table_result_MDA(ABS_file, protein_file)[0][2]
    sheet["F24"] = table_result_MDA(ABS_file, protein_file)[0][3]
    sheet["G24"] = table_result_MDA(ABS_file, protein_file)[0][4]
    sheet["H24"] = table_result_MDA(ABS_file, protein_file)[0][5]
    sheet["I24"] = table_result_MDA(ABS_file, protein_file)[0][6]
    sheet["J24"] = table_result_MDA(ABS_file, protein_file)[0][7]
    sheet["K24"] = table_result_MDA(ABS_file, protein_file)[0][8]
    sheet["L24"] = table_result_MDA(ABS_file, protein_file)[0][9]
    sheet["M24"] = table_result_MDA(ABS_file, protein_file)[0][10]
    sheet["N24"] = table_result_MDA(ABS_file, protein_file)[0][11]

    sheet["C25"] = table_result_MDA(ABS_file, protein_file)[1][0]
    sheet["D25"] = table_result_MDA(ABS_file, protein_file)[1][1]
    sheet["E25"] = table_result_MDA(ABS_file, protein_file)[1][2]
    sheet["F25"] = table_result_MDA(ABS_file, protein_file)[1][3]
    sheet["G25"] = table_result_MDA(ABS_file, protein_file)[1][4]
    sheet["H25"] = table_result_MDA(ABS_file, protein_file)[1][5]
    sheet["I25"] = table_result_MDA(ABS_file, protein_file)[1][6]
    sheet["J25"] = table_result_MDA(ABS_file, protein_file)[1][7]
    sheet["K25"] = table_result_MDA(ABS_file, protein_file)[1][8]
    sheet["L25"] = table_result_MDA(ABS_file, protein_file)[1][9]
    sheet["M25"] = table_result_MDA(ABS_file, protein_file)[1][10]
    sheet["N25"] = table_result_MDA(ABS_file, protein_file)[1][11]

    sheet["C26"] = table_result_MDA(ABS_file, protein_file)[2][0]
    sheet["D26"] = table_result_MDA(ABS_file, protein_file)[2][1]
    sheet["E26"] = table_result_MDA(ABS_file, protein_file)[2][2]
    sheet["F26"] = table_result_MDA(ABS_file, protein_file)[2][3]
    sheet["G26"] = table_result_MDA(ABS_file, protein_file)[2][4]
    sheet["H26"] = table_result_MDA(ABS_file, protein_file)[2][5]
    sheet["I26"] = table_result_MDA(ABS_file, protein_file)[2][6]
    sheet["J26"] = table_result_MDA(ABS_file, protein_file)[2][7]
    sheet["K26"] = table_result_MDA(ABS_file, protein_file)[2][8]
    sheet["L26"] = table_result_MDA(ABS_file, protein_file)[2][9]
    sheet["M26"] = table_result_MDA(ABS_file, protein_file)[2][10]
    sheet["N26"] = table_result_MDA(ABS_file, protein_file)[2][11]

    sheet["C27"] = table_result_MDA(ABS_file, protein_file)[3][0]
    sheet["D27"] = table_result_MDA(ABS_file, protein_file)[3][1]
    sheet["E27"] = table_result_MDA(ABS_file, protein_file)[3][2]
    sheet["F27"] = table_result_MDA(ABS_file, protein_file)[3][3]
    sheet["G27"] = table_result_MDA(ABS_file, protein_file)[3][4]
    sheet["H27"] = table_result_MDA(ABS_file, protein_file)[3][5]
    sheet["I27"] = table_result_MDA(ABS_file, protein_file)[3][6]
    sheet["J27"] = table_result_MDA(ABS_file, protein_file)[3][7]
    sheet["K27"] = table_result_MDA(ABS_file, protein_file)[3][8]
    sheet["L27"] = table_result_MDA(ABS_file, protein_file)[3][9]
    sheet["M27"] = table_result_MDA(ABS_file, protein_file)[3][10]
    sheet["N27"] = table_result_MDA(ABS_file, protein_file)[3][11]

    sheet["C28"] = table_result_MDA(ABS_file, protein_file)[4][0]
    sheet["D28"] = table_result_MDA(ABS_file, protein_file)[4][1]
    sheet["E28"] = table_result_MDA(ABS_file, protein_file)[4][2]
    sheet["F28"] = table_result_MDA(ABS_file, protein_file)[4][3]
    sheet["G28"] = table_result_MDA(ABS_file, protein_file)[4][4]
    sheet["H28"] = table_result_MDA(ABS_file, protein_file)[4][5]
    sheet["I28"] = table_result_MDA(ABS_file, protein_file)[4][6]
    sheet["J28"] = table_result_MDA(ABS_file, protein_file)[4][7]
    sheet["K28"] = table_result_MDA(ABS_file, protein_file)[4][8]
    sheet["L28"] = table_result_MDA(ABS_file, protein_file)[4][9]
    sheet["M28"] = table_result_MDA(ABS_file, protein_file)[4][10]
    sheet["N28"] = table_result_MDA(ABS_file, protein_file)[4][11]

    sheet["C29"] = table_result_MDA(ABS_file, protein_file)[5][0]
    sheet["D29"] = table_result_MDA(ABS_file, protein_file)[5][1]
    sheet["E29"] = table_result_MDA(ABS_file, protein_file)[5][2]
    sheet["F29"] = table_result_MDA(ABS_file, protein_file)[5][3]
    sheet["G29"] = table_result_MDA(ABS_file, protein_file)[5][4]
    sheet["H29"] = table_result_MDA(ABS_file, protein_file)[5][5]
    sheet["I29"] = table_result_MDA(ABS_file, protein_file)[5][6]
    sheet["J29"] = table_result_MDA(ABS_file, protein_file)[5][7]
    sheet["K29"] = table_result_MDA(ABS_file, protein_file)[5][8]
    sheet["L29"] = table_result_MDA(ABS_file, protein_file)[5][9]
    sheet["M29"] = table_result_MDA(ABS_file, protein_file)[5][10]
    sheet["N29"] = table_result_MDA(ABS_file, protein_file)[5][11]

    sheet["C30"] = table_result_MDA(ABS_file, protein_file)[6][0]
    sheet["D30"] = table_result_MDA(ABS_file, protein_file)[6][1]
    sheet["E30"] = table_result_MDA(ABS_file, protein_file)[6][2]
    sheet["F30"] = table_result_MDA(ABS_file, protein_file)[6][3]
    sheet["G30"] = table_result_MDA(ABS_file, protein_file)[6][4]
    sheet["H30"] = table_result_MDA(ABS_file, protein_file)[6][5]
    sheet["I30"] = table_result_MDA(ABS_file, protein_file)[6][6]
    sheet["J30"] = table_result_MDA(ABS_file, protein_file)[6][7]
    sheet["K30"] = table_result_MDA(ABS_file, protein_file)[6][8]
    sheet["L30"] = table_result_MDA(ABS_file, protein_file)[6][9]
    sheet["M30"] = table_result_MDA(ABS_file, protein_file)[6][10]
    sheet["N30"] = table_result_MDA(ABS_file, protein_file)[6][11]

    sheet["C31"] = table_result_MDA(ABS_file, protein_file)[7][0]
    sheet["D31"] = table_result_MDA(ABS_file, protein_file)[7][1]
    sheet["E31"] = table_result_MDA(ABS_file, protein_file)[7][2]
    sheet["F31"] = table_result_MDA(ABS_file, protein_file)[7][3]
    sheet["G31"] = table_result_MDA(ABS_file, protein_file)[7][4]
    sheet["H31"] = table_result_MDA(ABS_file, protein_file)[7][5]
    sheet["I31"] = table_result_MDA(ABS_file, protein_file)[7][6]
    sheet["J31"] = table_result_MDA(ABS_file, protein_file)[7][7]
    sheet["K31"] = table_result_MDA(ABS_file, protein_file)[7][8]
    sheet["L31"] = table_result_MDA(ABS_file, protein_file)[7][9]
    sheet["M31"] = table_result_MDA(ABS_file, protein_file)[7][10]
    sheet["N31"] = table_result_MDA(ABS_file, protein_file)[7][11]


    values = Reference(sheet, min_col=3, min_row=24,
                       max_col=14, max_row=31)
    chart = BarChart()
    chart.add_data(values)
    chart.title = " BAR-CHART "
    chart.x_axis.title = " ABS values"
    chart.y_axis.title = " PLATES "
    sheet.add_chart(chart, "B6")

    workbook.save(filename=(result+".xlsx"))


def selection_1():
    ABS_file = input("Enter file containing ABS values: ")
    result = input("Save as: ")
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    from openpyxl.chart import BarChart, Reference

    sheet["B24"] = "A"
    sheet["B25"] = "B"
    sheet["B26"] = "C"
    sheet["B27"] = "D"
    sheet["B28"] = "E"
    sheet["B29"] = "F"
    sheet["B30"] = "G"
    sheet["B31"] = "H"
    sheet["C23"] = "1"
    sheet["D23"] = "2"
    sheet["E23"] = "3"
    sheet["F23"] = "4"
    sheet["G23"] = "5"
    sheet["H23"] = "6"
    sheet["I23"] = "7"
    sheet["J23"] = "8"
    sheet["K23"] = "9"
    sheet["L23"] = "10"
    sheet["M23"] = "11"
    sheet["N23"] = "12"

    sheet["C24"] = table_result_sulphy(ABS_file)[0][0]
    sheet["D24"] = table_result_sulphy(ABS_file)[0][1]
    sheet["E24"] = table_result_sulphy(ABS_file)[0][2]
    sheet["F24"] = table_result_sulphy(ABS_file)[0][3]
    sheet["G24"] = table_result_sulphy(ABS_file)[0][4]
    sheet["H24"] = table_result_sulphy(ABS_file)[0][5]
    sheet["I24"] = table_result_sulphy(ABS_file)[0][6]
    sheet["J24"] = table_result_sulphy(ABS_file)[0][7]
    sheet["K24"] = table_result_sulphy(ABS_file)[0][8]
    sheet["L24"] = table_result_sulphy(ABS_file)[0][9]
    sheet["M24"] = table_result_sulphy(ABS_file)[0][10]
    sheet["N24"] = table_result_sulphy(ABS_file)[0][11]

    sheet["C25"] = table_result_sulphy(ABS_file)[1][0]
    sheet["D25"] = table_result_sulphy(ABS_file)[1][1]
    sheet["E25"] = table_result_sulphy(ABS_file)[1][2]
    sheet["F25"] = table_result_sulphy(ABS_file)[1][3]
    sheet["G25"] = table_result_sulphy(ABS_file)[1][4]
    sheet["H25"] = table_result_sulphy(ABS_file)[1][5]
    sheet["I25"] = table_result_sulphy(ABS_file)[1][6]
    sheet["J25"] = table_result_sulphy(ABS_file)[1][7]
    sheet["K25"] = table_result_sulphy(ABS_file)[1][8]
    sheet["L25"] = table_result_sulphy(ABS_file)[1][9]
    sheet["M25"] = table_result_sulphy(ABS_file)[1][10]
    sheet["N25"] = table_result_sulphy(ABS_file)[1][11]

    sheet["C26"] = table_result_sulphy(ABS_file)[2][0]
    sheet["D26"] = table_result_sulphy(ABS_file)[2][1]
    sheet["E26"] = table_result_sulphy(ABS_file)[2][2]
    sheet["F26"] = table_result_sulphy(ABS_file)[2][3]
    sheet["G26"] = table_result_sulphy(ABS_file)[2][4]
    sheet["H26"] = table_result_sulphy(ABS_file)[2][5]
    sheet["I26"] = table_result_sulphy(ABS_file)[2][6]
    sheet["J26"] = table_result_sulphy(ABS_file)[2][7]
    sheet["K26"] = table_result_sulphy(ABS_file)[2][8]
    sheet["L26"] = table_result_sulphy(ABS_file)[2][9]
    sheet["M26"] = table_result_sulphy(ABS_file)[2][10]
    sheet["N26"] = table_result_sulphy(ABS_file)[2][11]

    sheet["C27"] = table_result_sulphy(ABS_file)[3][0]
    sheet["D27"] = table_result_sulphy(ABS_file)[3][1]
    sheet["E27"] = table_result_sulphy(ABS_file)[3][2]
    sheet["F27"] = table_result_sulphy(ABS_file)[3][3]
    sheet["G27"] = table_result_sulphy(ABS_file)[3][4]
    sheet["H27"] = table_result_sulphy(ABS_file)[3][5]
    sheet["I27"] = table_result_sulphy(ABS_file)[3][6]
    sheet["J27"] = table_result_sulphy(ABS_file)[3][7]
    sheet["K27"] = table_result_sulphy(ABS_file)[3][8]
    sheet["L27"] = table_result_sulphy(ABS_file)[3][9]
    sheet["M27"] = table_result_sulphy(ABS_file)[3][10]
    sheet["N27"] = table_result_sulphy(ABS_file)[3][11]

    sheet["C28"] = table_result_sulphy(ABS_file)[4][0]
    sheet["D28"] = table_result_sulphy(ABS_file)[4][1]
    sheet["E28"] = table_result_sulphy(ABS_file)[4][2]
    sheet["F28"] = table_result_sulphy(ABS_file)[4][3]
    sheet["G28"] = table_result_sulphy(ABS_file)[4][4]
    sheet["H28"] = table_result_sulphy(ABS_file)[4][5]
    sheet["I28"] = table_result_sulphy(ABS_file)[4][6]
    sheet["J28"] = table_result_sulphy(ABS_file)[4][7]
    sheet["K28"] = table_result_sulphy(ABS_file)[4][8]
    sheet["L28"] = table_result_sulphy(ABS_file)[4][9]
    sheet["M28"] = table_result_sulphy(ABS_file)[4][10]
    sheet["N28"] = table_result_sulphy(ABS_file)[4][11]

    sheet["C29"] = table_result_sulphy(ABS_file)[5][0]
    sheet["D29"] = table_result_sulphy(ABS_file)[5][1]
    sheet["E29"] = table_result_sulphy(ABS_file)[5][2]
    sheet["F29"] = table_result_sulphy(ABS_file)[5][3]
    sheet["G29"] = table_result_sulphy(ABS_file)[5][4]
    sheet["H29"] = table_result_sulphy(ABS_file)[5][5]
    sheet["I29"] = table_result_sulphy(ABS_file)[5][6]
    sheet["J29"] = table_result_sulphy(ABS_file)[5][7]
    sheet["K29"] = table_result_sulphy(ABS_file)[5][8]
    sheet["L29"] = table_result_sulphy(ABS_file)[5][9]
    sheet["M29"] = table_result_sulphy(ABS_file)[5][10]
    sheet["N29"] = table_result_sulphy(ABS_file)[5][11]

    sheet["C30"] = table_result_sulphy(ABS_file)[6][0]
    sheet["D30"] = table_result_sulphy(ABS_file)[6][1]
    sheet["E30"] = table_result_sulphy(ABS_file)[6][2]
    sheet["F30"] = table_result_sulphy(ABS_file)[6][3]
    sheet["G30"] = table_result_sulphy(ABS_file)[6][4]
    sheet["H30"] = table_result_sulphy(ABS_file)[6][5]
    sheet["I30"] = table_result_sulphy(ABS_file)[6][6]
    sheet["J30"] = table_result_sulphy(ABS_file)[6][7]
    sheet["K30"] = table_result_sulphy(ABS_file)[6][8]
    sheet["L30"] = table_result_sulphy(ABS_file)[6][9]
    sheet["M30"] = table_result_sulphy(ABS_file)[6][10]
    sheet["N30"] = table_result_sulphy(ABS_file)[6][11]

    sheet["C31"] = table_result_sulphy(ABS_file)[7][0]
    sheet["D31"] = table_result_sulphy(ABS_file)[7][1]
    sheet["E31"] = table_result_sulphy(ABS_file)[7][2]
    sheet["F31"] = table_result_sulphy(ABS_file)[7][3]
    sheet["G31"] = table_result_sulphy(ABS_file)[7][4]
    sheet["H31"] = table_result_sulphy(ABS_file)[7][5]
    sheet["I31"] = table_result_sulphy(ABS_file)[7][6]
    sheet["J31"] = table_result_sulphy(ABS_file)[7][7]
    sheet["K31"] = table_result_sulphy(ABS_file)[7][8]
    sheet["L31"] = table_result_sulphy(ABS_file)[7][9]
    sheet["M31"] = table_result_sulphy(ABS_file)[7][10]
    sheet["N31"] = table_result_sulphy(ABS_file)[7][11]

    values = Reference(sheet, min_col=3, min_row=24,
                       max_col=14, max_row=31)
    chart = BarChart()
    chart.add_data(values)
    chart.title = " BAR-CHART "
    chart.x_axis.title = " ABS values"
    chart.y_axis.title = " PLATES "
    sheet.add_chart(chart, "B6")


    workbook.save(filename="Result.xlsx")


def selection_5():
    ABS_file = input("Enter file containing ABS values: ")
    protein_file = input("Enter file containing Protein values: ")
    result = input("Save as: ")
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    from openpyxl.chart import BarChart, Reference


    sheet["B24"] = "A"
    sheet["B25"] = "B"
    sheet["B26"] = "C"
    sheet["B27"] = "D"
    sheet["B28"] = "E"
    sheet["B29"] = "F"
    sheet["B30"] = "G"
    sheet["B31"] = "H"
    sheet["C23"] = "1"
    sheet["D23"] = "2"
    sheet["E23"] = "3"
    sheet["F23"] = "4"
    sheet["G23"] = "5"
    sheet["H23"] = "6"
    sheet["I23"] = "7"
    sheet["J23"] = "8"
    sheet["K23"] = "9"
    sheet["L23"] = "10"
    sheet["M23"] = "11"
    sheet["N23"] = "12"

    sheet["C24"] = table_result_H2O2(ABS_file, protein_file)[0][0]
    sheet["D24"] = table_result_H2O2(ABS_file, protein_file)[0][1]
    sheet["E24"] = table_result_H2O2(ABS_file, protein_file)[0][2]
    sheet["F24"] = table_result_H2O2(ABS_file, protein_file)[0][3]
    sheet["G24"] = table_result_H2O2(ABS_file, protein_file)[0][4]
    sheet["H24"] = table_result_H2O2(ABS_file, protein_file)[0][5]
    sheet["I24"] = table_result_H2O2(ABS_file, protein_file)[0][6]
    sheet["J24"] = table_result_H2O2(ABS_file, protein_file)[0][7]
    sheet["K24"] = table_result_H2O2(ABS_file, protein_file)[0][8]
    sheet["L24"] = table_result_H2O2(ABS_file, protein_file)[0][9]
    sheet["M24"] = table_result_H2O2(ABS_file, protein_file)[0][10]
    sheet["N24"] = table_result_H2O2(ABS_file, protein_file)[0][11]

    sheet["C25"] = table_result_H2O2(ABS_file, protein_file)[1][0]
    sheet["D25"] = table_result_H2O2(ABS_file, protein_file)[1][1]
    sheet["E25"] = table_result_H2O2(ABS_file, protein_file)[1][2]
    sheet["F25"] = table_result_H2O2(ABS_file, protein_file)[1][3]
    sheet["G25"] = table_result_H2O2(ABS_file, protein_file)[1][4]
    sheet["H25"] = table_result_H2O2(ABS_file, protein_file)[1][5]
    sheet["I25"] = table_result_H2O2(ABS_file, protein_file)[1][6]
    sheet["J25"] = table_result_H2O2(ABS_file, protein_file)[1][7]
    sheet["K25"] = table_result_H2O2(ABS_file, protein_file)[1][8]
    sheet["L25"] = table_result_H2O2(ABS_file, protein_file)[1][9]
    sheet["M25"] = table_result_H2O2(ABS_file, protein_file)[1][10]
    sheet["N25"] = table_result_H2O2(ABS_file, protein_file)[1][11]

    sheet["C26"] = table_result_H2O2(ABS_file, protein_file)[2][0]
    sheet["D26"] = table_result_H2O2(ABS_file, protein_file)[2][1]
    sheet["E26"] = table_result_H2O2(ABS_file, protein_file)[2][2]
    sheet["F26"] = table_result_H2O2(ABS_file, protein_file)[2][3]
    sheet["G26"] = table_result_H2O2(ABS_file, protein_file)[2][4]
    sheet["H26"] = table_result_H2O2(ABS_file, protein_file)[2][5]
    sheet["I26"] = table_result_H2O2(ABS_file, protein_file)[2][6]
    sheet["J26"] = table_result_H2O2(ABS_file, protein_file)[2][7]
    sheet["K26"] = table_result_H2O2(ABS_file, protein_file)[2][8]
    sheet["L26"] = table_result_H2O2(ABS_file, protein_file)[2][9]
    sheet["M26"] = table_result_H2O2(ABS_file, protein_file)[2][10]
    sheet["N26"] = table_result_H2O2(ABS_file, protein_file)[2][11]

    sheet["C27"] = table_result_H2O2(ABS_file, protein_file)[3][0]
    sheet["D27"] = table_result_H2O2(ABS_file, protein_file)[3][1]
    sheet["E27"] = table_result_H2O2(ABS_file, protein_file)[3][2]
    sheet["F27"] = table_result_H2O2(ABS_file, protein_file)[3][3]
    sheet["G27"] = table_result_H2O2(ABS_file, protein_file)[3][4]
    sheet["H27"] = table_result_H2O2(ABS_file, protein_file)[3][5]
    sheet["I27"] = table_result_H2O2(ABS_file, protein_file)[3][6]
    sheet["J27"] = table_result_H2O2(ABS_file, protein_file)[3][7]
    sheet["K27"] = table_result_H2O2(ABS_file, protein_file)[3][8]
    sheet["L27"] = table_result_H2O2(ABS_file, protein_file)[3][9]
    sheet["M27"] = table_result_H2O2(ABS_file, protein_file)[3][10]
    sheet["N27"] = table_result_H2O2(ABS_file, protein_file)[3][11]

    sheet["C28"] = table_result_H2O2(ABS_file, protein_file)[4][0]
    sheet["D28"] = table_result_H2O2(ABS_file, protein_file)[4][1]
    sheet["E28"] = table_result_H2O2(ABS_file, protein_file)[4][2]
    sheet["F28"] = table_result_H2O2(ABS_file, protein_file)[4][3]
    sheet["G28"] = table_result_H2O2(ABS_file, protein_file)[4][4]
    sheet["H28"] = table_result_H2O2(ABS_file, protein_file)[4][5]
    sheet["I28"] = table_result_H2O2(ABS_file, protein_file)[4][6]
    sheet["J28"] = table_result_H2O2(ABS_file, protein_file)[4][7]
    sheet["K28"] = table_result_H2O2(ABS_file, protein_file)[4][8]
    sheet["L28"] = table_result_H2O2(ABS_file, protein_file)[4][9]
    sheet["M28"] = table_result_H2O2(ABS_file, protein_file)[4][10]
    sheet["N28"] = table_result_H2O2(ABS_file, protein_file)[4][11]

    sheet["C27"] = table_result_H2O2(ABS_file, protein_file)[5][0]
    sheet["D29"] = table_result_H2O2(ABS_file, protein_file)[5][1]
    sheet["E29"] = table_result_H2O2(ABS_file, protein_file)[5][2]
    sheet["F29"] = table_result_H2O2(ABS_file, protein_file)[5][3]
    sheet["G29"] = table_result_H2O2(ABS_file, protein_file)[5][4]
    sheet["H29"] = table_result_H2O2(ABS_file, protein_file)[5][5]
    sheet["I29"] = table_result_H2O2(ABS_file, protein_file)[5][6]
    sheet["J29"] = table_result_H2O2(ABS_file, protein_file)[5][7]
    sheet["K29"] = table_result_H2O2(ABS_file, protein_file)[5][8]
    sheet["L29"] = table_result_H2O2(ABS_file, protein_file)[5][9]
    sheet["M29"] = table_result_H2O2(ABS_file, protein_file)[5][10]
    sheet["N29"] = table_result_H2O2(ABS_file, protein_file)[5][11]

    sheet["C30"] = table_result_H2O2(ABS_file, protein_file)[6][0]
    sheet["D30"] = table_result_H2O2(ABS_file, protein_file)[6][1]
    sheet["E30"] = table_result_H2O2(ABS_file, protein_file)[6][2]
    sheet["F30"] = table_result_H2O2(ABS_file, protein_file)[6][3]
    sheet["G30"] = table_result_H2O2(ABS_file, protein_file)[6][4]
    sheet["H30"] = table_result_H2O2(ABS_file, protein_file)[6][5]
    sheet["I30"] = table_result_H2O2(ABS_file, protein_file)[6][6]
    sheet["J30"] = table_result_H2O2(ABS_file, protein_file)[6][7]
    sheet["K30"] = table_result_H2O2(ABS_file, protein_file)[6][8]
    sheet["L30"] = table_result_H2O2(ABS_file, protein_file)[6][9]
    sheet["M30"] = table_result_H2O2(ABS_file, protein_file)[6][10]
    sheet["N30"] = table_result_H2O2(ABS_file, protein_file)[6][11]

    sheet["C31"] = table_result_H2O2(ABS_file, protein_file)[7][0]
    sheet["D31"] = table_result_H2O2(ABS_file, protein_file)[7][1]
    sheet["E31"] = table_result_H2O2(ABS_file, protein_file)[7][2]
    sheet["F31"] = table_result_H2O2(ABS_file, protein_file)[7][3]
    sheet["G31"] = table_result_H2O2(ABS_file, protein_file)[7][4]
    sheet["H31"] = table_result_H2O2(ABS_file, protein_file)[7][5]
    sheet["I31"] = table_result_H2O2(ABS_file, protein_file)[7][6]
    sheet["J31"] = table_result_H2O2(ABS_file, protein_file)[7][7]
    sheet["K31"] = table_result_H2O2(ABS_file, protein_file)[7][8]
    sheet["L31"] = table_result_H2O2(ABS_file, protein_file)[7][9]
    sheet["M31"] = table_result_H2O2(ABS_file, protein_file)[7][10]
    sheet["N31"] = table_result_H2O2(ABS_file, protein_file)[7][11]

    values = Reference(sheet, min_col=3, min_row=24,
                       max_col=14, max_row=31)
    chart = BarChart()
    chart.add_data(values)
    chart.title = " BAR-CHART "
    chart.x_axis.title = " ABS values"
    chart.y_axis.title = " PLATES "
    sheet.add_chart(chart, "B6")



    workbook.save(filename=(result+".xlsx"))


def selection_2():
    ABS_file = input("Enter file containing ABS values: ")
    protein_file = input("Enter file containing protein values: ")
    result = input("Save as: ")
    from openpyxl.chart import BarChart, Reference

    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active

    sheet["B24"] = "A"
    sheet["B25"] = "B"
    sheet["B26"] = "C"
    sheet["B27"] = "D"
    sheet["B28"] = "E"
    sheet["B29"] = "F"
    sheet["B30"] = "G"
    sheet["B31"] = "H"
    sheet["C23"] = "1"
    sheet["D23"] = "2"
    sheet["E23"] = "3"
    sheet["F23"] = "4"
    sheet["G23"] = "5"
    sheet["H23"] = "6"
    sheet["I23"] = "7"
    sheet["J23"] = "8"
    sheet["K23"] = "9"
    sheet["L23"] = "10"
    sheet["M23"] = "11"
    sheet["N23"] = "12"

    sheet["C24"] = table_result_Atpase(ABS_file, protein_file)[0][0]
    sheet["D24"] = table_result_Atpase(ABS_file, protein_file)[0][1]
    sheet["E24"] = table_result_Atpase(ABS_file, protein_file)[0][2]
    sheet["F24"] = table_result_Atpase(ABS_file, protein_file)[0][3]
    sheet["G24"] = table_result_Atpase(ABS_file, protein_file)[0][4]
    sheet["H24"] = table_result_Atpase(ABS_file, protein_file)[0][5]
    sheet["I24"] = table_result_Atpase(ABS_file, protein_file)[0][6]
    sheet["J24"] = table_result_Atpase(ABS_file, protein_file)[0][7]
    sheet["K24"] = table_result_Atpase(ABS_file, protein_file)[0][8]
    sheet["L24"] = table_result_Atpase(ABS_file, protein_file)[0][9]
    sheet["M24"] = table_result_Atpase(ABS_file, protein_file)[0][10]
    sheet["N24"] = table_result_Atpase(ABS_file, protein_file)[0][11]

    sheet["C25"] = table_result_Atpase(ABS_file, protein_file)[1][0]
    sheet["D25"] = table_result_Atpase(ABS_file, protein_file)[1][1]
    sheet["E25"] = table_result_Atpase(ABS_file, protein_file)[1][2]
    sheet["F25"] = table_result_Atpase(ABS_file, protein_file)[1][3]
    sheet["G25"] = table_result_Atpase(ABS_file, protein_file)[1][4]
    sheet["H25"] = table_result_Atpase(ABS_file, protein_file)[1][5]
    sheet["I25"] = table_result_Atpase(ABS_file, protein_file)[1][6]
    sheet["J25"] = table_result_Atpase(ABS_file, protein_file)[1][7]
    sheet["K25"] = table_result_Atpase(ABS_file, protein_file)[1][8]
    sheet["L25"] = table_result_Atpase(ABS_file, protein_file)[1][9]
    sheet["M25"] = table_result_Atpase(ABS_file, protein_file)[1][10]
    sheet["N25"] = table_result_Atpase(ABS_file, protein_file)[1][11]

    sheet["C26"] = table_result_Atpase(ABS_file, protein_file)[2][0]
    sheet["D26"] = table_result_Atpase(ABS_file, protein_file)[2][1]
    sheet["E26"] = table_result_Atpase(ABS_file, protein_file)[2][2]
    sheet["F26"] = table_result_Atpase(ABS_file, protein_file)[2][3]
    sheet["G26"] = table_result_Atpase(ABS_file, protein_file)[2][4]
    sheet["H26"] = table_result_Atpase(ABS_file, protein_file)[2][5]
    sheet["I26"] = table_result_Atpase(ABS_file, protein_file)[2][6]
    sheet["J26"] = table_result_Atpase(ABS_file, protein_file)[2][7]
    sheet["K26"] = table_result_Atpase(ABS_file, protein_file)[2][8]
    sheet["L26"] = table_result_Atpase(ABS_file, protein_file)[2][9]
    sheet["M26"] = table_result_Atpase(ABS_file, protein_file)[2][10]
    sheet["N26"] = table_result_Atpase(ABS_file, protein_file)[2][11]

    sheet["C27"] = table_result_Atpase(ABS_file, protein_file)[3][0]
    sheet["D27"] = table_result_Atpase(ABS_file, protein_file)[3][1]
    sheet["E27"] = table_result_Atpase(ABS_file, protein_file)[3][2]
    sheet["F27"] = table_result_Atpase(ABS_file, protein_file)[3][3]
    sheet["G27"] = table_result_Atpase(ABS_file, protein_file)[3][4]
    sheet["H27"] = table_result_Atpase(ABS_file, protein_file)[3][5]
    sheet["I27"] = table_result_Atpase(ABS_file, protein_file)[3][6]
    sheet["J27"] = table_result_Atpase(ABS_file, protein_file)[3][7]
    sheet["K27"] = table_result_Atpase(ABS_file, protein_file)[3][8]
    sheet["L27"] = table_result_Atpase(ABS_file, protein_file)[3][9]
    sheet["M27"] = table_result_Atpase(ABS_file, protein_file)[3][10]
    sheet["N27"] = table_result_Atpase(ABS_file, protein_file)[3][11]

    sheet["C28"] = table_result_Atpase(ABS_file, protein_file)[4][0]
    sheet["D28"] = table_result_Atpase(ABS_file, protein_file)[4][1]
    sheet["E28"] = table_result_Atpase(ABS_file, protein_file)[4][2]
    sheet["F28"] = table_result_Atpase(ABS_file, protein_file)[4][3]
    sheet["G28"] = table_result_Atpase(ABS_file, protein_file)[4][4]
    sheet["H28"] = table_result_Atpase(ABS_file, protein_file)[4][5]
    sheet["I28"] = table_result_Atpase(ABS_file, protein_file)[4][6]
    sheet["J28"] = table_result_Atpase(ABS_file, protein_file)[4][7]
    sheet["K28"] = table_result_Atpase(ABS_file, protein_file)[4][8]
    sheet["L28"] = table_result_Atpase(ABS_file, protein_file)[4][9]
    sheet["M28"] = table_result_Atpase(ABS_file, protein_file)[4][10]
    sheet["N28"] = table_result_Atpase(ABS_file, protein_file)[4][11]

    sheet["C29"] = table_result_Atpase(ABS_file, protein_file)[5][0]
    sheet["D29"] = table_result_Atpase(ABS_file, protein_file)[5][1]
    sheet["E29"] = table_result_Atpase(ABS_file, protein_file)[5][2]
    sheet["F29"] = table_result_Atpase(ABS_file, protein_file)[5][3]
    sheet["G29"] = table_result_Atpase(ABS_file, protein_file)[5][4]
    sheet["H29"] = table_result_Atpase(ABS_file, protein_file)[5][5]
    sheet["I29"] = table_result_Atpase(ABS_file, protein_file)[5][6]
    sheet["J29"] = table_result_Atpase(ABS_file, protein_file)[5][7]
    sheet["K29"] = table_result_Atpase(ABS_file, protein_file)[5][8]
    sheet["L29"] = table_result_Atpase(ABS_file, protein_file)[5][9]
    sheet["M29"] = table_result_Atpase(ABS_file, protein_file)[5][10]
    sheet["N29"] = table_result_Atpase(ABS_file, protein_file)[5][11]

    sheet["C30"] = table_result_Atpase(ABS_file, protein_file)[6][0]
    sheet["D30"] = table_result_Atpase(ABS_file, protein_file)[6][1]
    sheet["E30"] = table_result_Atpase(ABS_file, protein_file)[6][2]
    sheet["F30"] = table_result_Atpase(ABS_file, protein_file)[6][3]
    sheet["G30"] = table_result_Atpase(ABS_file, protein_file)[6][4]
    sheet["H30"] = table_result_Atpase(ABS_file, protein_file)[6][5]
    sheet["I30"] = table_result_Atpase(ABS_file, protein_file)[6][6]
    sheet["J30"] = table_result_Atpase(ABS_file, protein_file)[6][7]
    sheet["K30"] = table_result_Atpase(ABS_file, protein_file)[6][8]
    sheet["L30"] = table_result_Atpase(ABS_file, protein_file)[6][9]
    sheet["M30"] = table_result_Atpase(ABS_file, protein_file)[6][10]
    sheet["N30"] = table_result_Atpase(ABS_file, protein_file)[6][11]

    sheet["C31"] = table_result_Atpase(ABS_file, protein_file)[7][0]
    sheet["D31"] = table_result_Atpase(ABS_file, protein_file)[7][1]
    sheet["E31"] = table_result_Atpase(ABS_file, protein_file)[7][2]
    sheet["F31"] = table_result_Atpase(ABS_file, protein_file)[7][3]
    sheet["G31"] = table_result_Atpase(ABS_file, protein_file)[7][4]
    sheet["H31"] = table_result_Atpase(ABS_file, protein_file)[7][5]
    sheet["I31"] = table_result_Atpase(ABS_file, protein_file)[7][6]
    sheet["J31"] = table_result_Atpase(ABS_file, protein_file)[7][7]
    sheet["K31"] = table_result_Atpase(ABS_file, protein_file)[7][8]
    sheet["L31"] = table_result_Atpase(ABS_file, protein_file)[7][9]
    sheet["M31"] = table_result_Atpase(ABS_file, protein_file)[7][10]
    sheet["N31"] = table_result_Atpase(ABS_file, protein_file)[7][11]

    values = Reference(sheet, min_col=3, min_row=24,
                       max_col=14, max_row=31)
    chart = BarChart()
    chart.add_data(values)
    chart.title = " BAR-CHART "
    chart.x_axis.title = " ABS values"
    chart.y_axis.title = " PLATES "
    sheet.add_chart(chart, "B6")


    workbook.save(filename=(result+".xlsx"))


def selection_3():
    ABS_file = input("Enter file containing ABS values: ")
    result = input("Save as: ")
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    from openpyxl.chart import BarChart, Reference

    sheet["B24"] = "A"
    sheet["B25"] = "B"
    sheet["B26"] = "C"
    sheet["B27"] = "D"
    sheet["B28"] = "E"
    sheet["B29"] = "F"
    sheet["B30"] = "G"
    sheet["B31"] = "H"
    sheet["C23"] = "1"
    sheet["D23"] = "2"
    sheet["E23"] = "3"
    sheet["F23"] = "4"
    sheet["G23"] = "5"
    sheet["H23"] = "6"
    sheet["I23"] = "7"
    sheet["J23"] = "8"
    sheet["K23"] = "9"
    sheet["L23"] = "10"
    sheet["M23"] = "11"
    sheet["N23"] = "12"

    sheet["C24"] = table_result_carbonyl(ABS_file)[0][0]
    sheet["D24"] = table_result_carbonyl(ABS_file)[0][1]
    sheet["E24"] = table_result_carbonyl(ABS_file)[0][2]
    sheet["F24"] = table_result_carbonyl(ABS_file)[0][3]
    sheet["G24"] = table_result_carbonyl(ABS_file)[0][4]
    sheet["H24"] = table_result_carbonyl(ABS_file)[0][5]
    sheet["I24"] = table_result_carbonyl(ABS_file)[0][6]
    sheet["J24"] = table_result_carbonyl(ABS_file)[0][7]
    sheet["K24"] = table_result_carbonyl(ABS_file)[0][8]
    sheet["L24"] = table_result_carbonyl(ABS_file)[0][9]
    sheet["M24"] = table_result_carbonyl(ABS_file)[0][10]
    sheet["N24"] = table_result_carbonyl(ABS_file)[0][11]

    sheet["C25"] = table_result_carbonyl(ABS_file)[1][0]
    sheet["D25"] = table_result_carbonyl(ABS_file)[1][1]
    sheet["E25"] = table_result_carbonyl(ABS_file)[1][2]
    sheet["F25"] = table_result_carbonyl(ABS_file)[1][3]
    sheet["G25"] = table_result_carbonyl(ABS_file)[1][4]
    sheet["H25"] = table_result_carbonyl(ABS_file)[1][5]
    sheet["I25"] = table_result_carbonyl(ABS_file)[1][6]
    sheet["J25"] = table_result_carbonyl(ABS_file)[1][7]
    sheet["K25"] = table_result_carbonyl(ABS_file)[1][8]
    sheet["L25"] = table_result_carbonyl(ABS_file)[1][9]
    sheet["M25"] = table_result_carbonyl(ABS_file)[1][10]
    sheet["N25"] = table_result_carbonyl(ABS_file)[1][11]

    sheet["C26"] = table_result_carbonyl(ABS_file)[2][0]
    sheet["D26"] = table_result_carbonyl(ABS_file)[2][1]
    sheet["E26"] = table_result_carbonyl(ABS_file)[2][2]
    sheet["F26"] = table_result_carbonyl(ABS_file)[2][3]
    sheet["G26"] = table_result_carbonyl(ABS_file)[2][4]
    sheet["H26"] = table_result_carbonyl(ABS_file)[2][5]
    sheet["I26"] = table_result_carbonyl(ABS_file)[2][6]
    sheet["J26"] = table_result_carbonyl(ABS_file)[2][7]
    sheet["K26"] = table_result_carbonyl(ABS_file)[2][8]
    sheet["L26"] = table_result_carbonyl(ABS_file)[2][9]
    sheet["M26"] = table_result_carbonyl(ABS_file)[2][10]
    sheet["N26"] = table_result_carbonyl(ABS_file)[2][11]

    sheet["C27"] = table_result_carbonyl(ABS_file)[3][0]
    sheet["D27"] = table_result_carbonyl(ABS_file)[3][1]
    sheet["E27"] = table_result_carbonyl(ABS_file)[3][2]
    sheet["F27"] = table_result_carbonyl(ABS_file)[3][3]
    sheet["G27"] = table_result_carbonyl(ABS_file)[3][4]
    sheet["H27"] = table_result_carbonyl(ABS_file)[3][5]
    sheet["I27"] = table_result_carbonyl(ABS_file)[3][6]
    sheet["J27"] = table_result_carbonyl(ABS_file)[3][7]
    sheet["K27"] = table_result_carbonyl(ABS_file)[3][8]
    sheet["L27"] = table_result_carbonyl(ABS_file)[3][9]
    sheet["M27"] = table_result_carbonyl(ABS_file)[3][10]
    sheet["N27"] = table_result_carbonyl(ABS_file)[3][11]

    sheet["C28"] = table_result_carbonyl(ABS_file)[4][0]
    sheet["D28"] = table_result_carbonyl(ABS_file)[4][1]
    sheet["E28"] = table_result_carbonyl(ABS_file)[4][2]
    sheet["F28"] = table_result_carbonyl(ABS_file)[4][3]
    sheet["G28"] = table_result_carbonyl(ABS_file)[4][4]
    sheet["H28"] = table_result_carbonyl(ABS_file)[4][5]
    sheet["I28"] = table_result_carbonyl(ABS_file)[4][6]
    sheet["J28"] = table_result_carbonyl(ABS_file)[4][7]
    sheet["K28"] = table_result_carbonyl(ABS_file)[4][8]
    sheet["L28"] = table_result_carbonyl(ABS_file)[4][9]
    sheet["M28"] = table_result_carbonyl(ABS_file)[4][10]
    sheet["N28"] = table_result_carbonyl(ABS_file)[4][11]

    sheet["C29"] = table_result_carbonyl(ABS_file)[5][0]
    sheet["D29"] = table_result_carbonyl(ABS_file)[5][1]
    sheet["E29"] = table_result_carbonyl(ABS_file)[5][2]
    sheet["F29"] = table_result_carbonyl(ABS_file)[5][3]
    sheet["G29"] = table_result_carbonyl(ABS_file)[5][4]
    sheet["H29"] = table_result_carbonyl(ABS_file)[5][5]
    sheet["I29"] = table_result_carbonyl(ABS_file)[5][6]
    sheet["J29"] = table_result_carbonyl(ABS_file)[5][7]
    sheet["K29"] = table_result_carbonyl(ABS_file)[5][8]
    sheet["L29"] = table_result_carbonyl(ABS_file)[5][9]
    sheet["M29"] = table_result_carbonyl(ABS_file)[5][10]
    sheet["N29"] = table_result_carbonyl(ABS_file)[5][11]

    sheet["C30"] = table_result_carbonyl(ABS_file)[6][0]
    sheet["D30"] = table_result_carbonyl(ABS_file)[6][1]
    sheet["E30"] = table_result_carbonyl(ABS_file)[6][2]
    sheet["F30"] = table_result_carbonyl(ABS_file)[6][3]
    sheet["G30"] = table_result_carbonyl(ABS_file)[6][4]
    sheet["H30"] = table_result_carbonyl(ABS_file)[6][5]
    sheet["I30"] = table_result_carbonyl(ABS_file)[6][6]
    sheet["J30"] = table_result_carbonyl(ABS_file)[6][7]
    sheet["K30"] = table_result_carbonyl(ABS_file)[6][8]
    sheet["L30"] = table_result_carbonyl(ABS_file)[6][9]
    sheet["M30"] = table_result_carbonyl(ABS_file)[6][10]
    sheet["N30"] = table_result_carbonyl(ABS_file)[6][11]

    sheet["C31"] = table_result_carbonyl(ABS_file)[7][0]
    sheet["D31"] = table_result_carbonyl(ABS_file)[7][1]
    sheet["E31"] = table_result_carbonyl(ABS_file)[7][2]
    sheet["F31"] = table_result_carbonyl(ABS_file)[7][3]
    sheet["G31"] = table_result_carbonyl(ABS_file)[7][4]
    sheet["H31"] = table_result_carbonyl(ABS_file)[7][5]
    sheet["I31"] = table_result_carbonyl(ABS_file)[7][6]
    sheet["J31"] = table_result_carbonyl(ABS_file)[7][7]
    sheet["K31"] = table_result_carbonyl(ABS_file)[7][8]
    sheet["L31"] = table_result_carbonyl(ABS_file)[7][9]
    sheet["M31"] = table_result_carbonyl(ABS_file)[7][10]
    sheet["N31"] = table_result_carbonyl(ABS_file)[7][11]

    values = Reference(sheet, min_col=3, min_row=24,
                       max_col=14, max_row=31)
    chart = BarChart()
    chart.add_data(values)
    chart.title = " BAR-CHART "
    chart.x_axis.title = " ABS values"
    chart.y_axis.title = " PLATES "
    sheet.add_chart(chart, "B6")

    workbook.save(filename=(result+".xlsx"))

