import multiprocessing
import time
start_time = time.time()


def read_rows(file_name="practice", rows=2, start_column=1, end_column=2):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=file_name)
    workbook.sheetnames
    sheet = workbook.active

    lists = []
    n = start_column - 1
    while (n <= (end_column)):
        n += 1
        cell = (sheet.cell(row=rows, column=n))
        lists.append(cell.value)
    return lists


def read_table(file_name="practice", start_row=1, end_row=2, start_column=1, end_column=2):
    lists = []
    n = start_row - 1
    end_row -= 1
    while (n <= (end_row)):
        n += 1
        lists.append(read_rows(file_name, n, start_column, end_column))
    return lists


def protein_values(file_protei="file"):
    file_protein = file_protei + ".xlsx"
    return (read_table(file_protein, 24, 31, 3, 13))


def ABS_values(file_ab="file"):
    file_abs = file_ab + ".xlsx"
    return (read_table(file_abs, 24, 31, 3, 13))


def Atpase(abs=0, protein=1):
    x = (abs / 0.2431) * 1000
    w = x / protein
    v = w / 0.2
    o = v / 60
    Atpase = o * 4.5
    return Atpase


def carbonyl(abs=0):
    car = (abs * 0.18) / 132000  # car = carbonyl
    return car


def MDA(abs=0, protein=1):
    a = abs * 3
    b = 1.56 * 100000 * 0.4 * protein
    MDA = a / b
    return MDA


def H2O2(abs=0, protein=1):
    a = 2.499 - abs
    b = 0.3175 * protein
    H2O2 = a / b
    return H2O2


def sulphy(abs=0):
    a = 1 - (abs / 14.150)
    b = (1.5 * a) / 1000
    c = (b * 1000) / 0.2
    return c


def column_result_Atpase(ABS_file="file", protein_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(Atpase((ABS_values()[r][n]), (protein_values()[r][n])))
        if (n == 11):
            break

    return lists


def table_result_Atpase(ABS_file="file", protein_file="file"):
    lists = []
    lists.append(column_result_Atpase(0))
    lists.append(column_result_Atpase(1))
    lists.append(column_result_Atpase(2))
    lists.append(column_result_Atpase(3))
    lists.append(column_result_Atpase(4))
    lists.append(column_result_Atpase(5))
    lists.append(column_result_Atpase(6))
    lists.append(column_result_Atpase(7))

    return lists


def column_result_sulphy(ABS_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(sulphy(ABS_values(ABS_file)[r][n]))
        if (n == 11):
            break

    return lists


def table_result_sulphy(ABS_file="file"):
    lists = []
    lists.append(column_result_sulphy(ABS_file, 0))
    lists.append(column_result_sulphy(ABS_file, 1))
    lists.append(column_result_sulphy(ABS_file, 2))
    lists.append(column_result_sulphy(ABS_file, 3))
    lists.append(column_result_sulphy(ABS_file, 4))
    lists.append(column_result_sulphy(ABS_file, 5))
    lists.append(column_result_sulphy(ABS_file, 6))
    lists.append(column_result_sulphy(ABS_file, 7))


def column_result_carbonyl(ABS_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(carbonyl(ABS_values(ABS_file)[r][n]))
        if (n == 11):
            break

    return lists


def table_result_carbonyl(ABS_file="file"):
    lists = []
    lists.append(column_result_carbonyl(ABS_file, 0))
    lists.append(column_result_carbonyl(ABS_file, 1))
    lists.append(column_result_carbonyl(ABS_file, 2))
    lists.append(column_result_carbonyl(ABS_file, 3))
    lists.append(column_result_carbonyl(ABS_file, 4))
    lists.append(column_result_carbonyl(ABS_file, 5))
    lists.append(column_result_carbonyl(ABS_file, 6))
    lists.append(column_result_carbonyl(ABS_file, 7))

    return lists


def column_result_H2O2(ABS_file="file", protein_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(H2O2((ABS_values(ABS_file)[r][n]), (protein_values(protein_file)[r][n])))
        if (n == 11):
            break

    return lists


def table_result_H2O2(ABS_file="file", protein_file="file"):
    lists = []
    lists.append(column_result_H2O2(ABS_file, protein_file, 0))
    lists.append(column_result_H2O2(ABS_file, protein_file, 1))
    lists.append(column_result_H2O2(ABS_file, protein_file, 2))
    lists.append(column_result_H2O2(ABS_file, protein_file, 3))
    lists.append(column_result_H2O2(ABS_file, protein_file, 4))
    lists.append(column_result_H2O2(ABS_file, protein_file, 5))
    lists.append(column_result_H2O2(ABS_file, protein_file, 6))
    lists.append(column_result_H2O2(ABS_file, protein_file, 7))

    return lists


def column_result_MDA(ABS_file="file", protein_file="file", r=0):
    lists = []
    n = -1
    while (n < 12):
        n += 1
        lists.append(MDA((ABS_values(ABS_file)[r][n]), (protein_values(protein_file)[r][n])))
        if (n == 11):
            break

    return lists


def table_result_MDA(ABS_file="file", protein_file="file"):
    lists = []
    lists.append(column_result_MDA(ABS_file, protein_file, 0))
    lists.append(column_result_MDA(ABS_file, protein_file, 1))
    lists.append(column_result_MDA(ABS_file, protein_file, 2))
    lists.append(column_result_MDA(ABS_file, protein_file, 3))
    lists.append(column_result_MDA(ABS_file, protein_file, 4))
    lists.append(column_result_MDA(ABS_file, protein_file, 5))
    lists.append(column_result_MDA(ABS_file, protein_file, 6))
    lists.append(column_result_MDA(ABS_file, protein_file, 7))

    return lists



select = int(input('''
Enter 1 for a "sulphy" table' 
Enter 2 for a "Carbonyl" table' 
Enter 3 for a "Atpase" table
Enter 4 for a "MDA" table
Enter 5 for a "H2O2" table
                '''))

if (select <= 2):
        ABS_file = input("Enter file containing ABS values: ")


else:
    if (select > 2):
        ABS_file = input("Enter file containing ABS values: ")
        protein_file = input("Enter file containing protein values: ")

def selection():
        if (select == 1):
            return table_result_sulphy(ABS_file)
        elif (select == 2):
            return table_result_carbonyl(ABS_file)

        elif (select == 3):
            return table_result_Atpase(ABS_file, protein_file)
        elif (select == 4):
            return table_result_MDA(ABS_file, protein_file)
        elif (select == 5):
            return table_result_H2O2(ABS_file, protein_file)




from openpyxl import Workbook
workbook = Workbook()
sheet = workbook.active
def result_1():
    sheet["C24"] = selection[0][0]
    sheet["D24"] = selection[0][1]
    sheet["E24"] = selection[0][2]
    sheet["F24"] = selection[0][3]
    sheet["G24"] = selection[0][4]
    sheet["H24"] = selection[0][5]
    sheet["I24"] = selection[0][6]
    sheet["J24"] = selection[0][7]
    sheet["K24"] = selection[0][8]
    sheet["L24"] = selection[0][9]
    sheet["M24"] = selection[0][10]
    sheet["N24"] = selection[0][11]

def result_2():
    sheet["C25"] = selection[1][0]
    sheet["D25"] = selection[1][1]
    sheet["E25"] = selection[1][2]
    sheet["F25"] = selection[1][3]
    sheet["G25"] = selection[1][4]
    sheet["H25"] = selection[1][5]
    sheet["I25"] = selection[1][6]
    sheet["J25"] = selection[1][7]
    sheet["K25"] = selection[1][8]
    sheet["L25"] = selection[1][9]
    sheet["M25"] = selection[1][10]
    sheet["N25"] = selection[1][11]

def result_3():
    sheet["C26"] = selection[2][0]
    sheet["D26"] = selection[2][1]
    sheet["E26"] = selection[2][2]
    sheet["F26"] = selection[2][3]
    sheet["G26"] = selection[2][4]
    sheet["H26"] = selection[2][5]
    sheet["I26"] = selection[2][6]
    sheet["J26"] = selection[2][7]
    sheet["K26"] = selection[2][8]
    sheet["L26"] = selection[2][9]
    sheet["M26"] = selection[2][10]
    sheet["N26"] = selection[2][11]

def result_4():
    sheet["C27"] = selection[3][0]
    sheet["D27"] = selection[3][1]
    sheet["E27"] = selection[3][2]
    sheet["F27"] = selection[3][3]
    sheet["G27"] = selection[3][4]
    sheet["H27"] = selection[3][5]
    sheet["I27"] = selection[3][6]
    sheet["J27"] = selection[3][7]
    sheet["K27"] = selection[3][8]
    sheet["L27"] = selection[3][9]
    sheet["M27"] = selection[3][10]
    sheet["N27"] = selection[3][11]

def result_5():
    sheet["C26"] = selection[4][0]
    sheet["D26"] = selection[4][1]
    sheet["E26"] = selection[4][2]
    sheet["F26"] = selection[4][3]
    sheet["G26"] = selection[4][4]
    sheet["H26"] = selection[4][5]
    sheet["I26"] = selection[4][6]
    sheet["J26"] = selection[4][7]
    sheet["K26"] = selection[4][8]
    sheet["L26"] = selection[4][9]
    sheet["M26"] = selection[4][10]
    sheet["N26"] = selection[4][11]
def result_6():
    sheet["C27"] = selection[5][0]
    sheet["D27"] = selection[5][1]
    sheet["E27"] = selection[5][2]
    sheet["F27"] = selection[5][3]
    sheet["G27"] = selection[5][4]
    sheet["H27"] = selection[5][5]
    sheet["I27"] = selection[5][6]
    sheet["J27"] = selection[5][7]
    sheet["K27"] = selection[5][8]
    sheet["L27"] = selection[5][9]
    sheet["M27"] = selection[5][10]
    sheet["N27"] = selection[5][11]


def result_7():

    sheet["C28"] = selection[6][0]
    sheet["D28"] = selection[6][1]
    sheet["E28"] = selection[6][2]
    sheet["F28"] = selection[6][3]
    sheet["G28"] = selection[6][4]
    sheet["H28"] = selection[6][5]
    sheet["I28"] = selection[6][6]
    sheet["J28"] = selection[6][7]
    sheet["K28"] = selection[6][8]
    sheet["L28"] = selection[6][9]
    sheet["M28"] = selection[6][10]
    sheet["N28"] = selection[6][11]
def result_8():
    sheet["C29"] = selection[7][0]
    sheet["D29"] = selection[7][1]
    sheet["E29"] = selection[7][2]
    sheet["F29"] = selection[7][3]
    sheet["G29"] = selection[7][4]
    sheet["H29"] = selection[7][5]
    sheet["I29"] = selection[7][6]
    sheet["J29"] = selection[7][7]
    sheet["K29"] = selection[7][8]
    sheet["L29"] = selection[7][9]
    sheet["M29"] = selection[7][10]
    sheet["N29"] = selection[7][11]



p1 = multiprocessing.Process(target=result_1())
p2 = multiprocessing.Process(target=result_2())
p3 = multiprocessing.Process(target=result_3())
p4 = multiprocessing.Process(target=result_4())
p5 = multiprocessing.Process(target=result_5())
p6 = multiprocessing.Process(target=result_6())
p7 = multiprocessing.Process(target=result_7())
p8 = multiprocessing.Process(target=result_8())

p1.start()
p2.start()
p3.start()
p4.start()
p5.start()
p6.start()
p7.start()
p8.start()

p1.join()
p2.join()
p3.join()
p4.join()
p5.join()
p6.join()
p7.join()
p8.join()


workbook.save(filename="result.xlsx")
print("--- %s seconds ---" % (time.time() - start_time))